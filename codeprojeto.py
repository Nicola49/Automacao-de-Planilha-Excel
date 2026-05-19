# percorrer toda a base de dados
# para cada item
#   ver se o bairro ja existe em uma aba, se não existir, criar tal aba
#   copiar os valores daquela linha e colocar na aba do bairro correspondente

from openpyxl import load_workbook
from copy import copy

def criar_aba(bairro, arquivo_base, estilo):
    if bairro not in arquivo_base.sheetnames:
        arquivo_base.create_sheet(bairro)
        nova_aba = arquivo_base[bairro]
        nova_aba["A1"].value = "Data de Nascimento"
        nova_aba["B1"].value = "Pessoa"
        nova_aba["C1"].value = "Bairro"
        nova_aba["A1"]._style = estilo
        nova_aba["B1"]._style = estilo
        nova_aba["C1"]._style = estilo

def transferir_info_aba(aba_origem, aba_destino, linha_origem):
    linha_destino = aba_destino.max_row + 1
    for coluna in range(1, 4):
        celula_origem = aba_origem.cell(row=linha_origem, column=coluna)
        celula_destino = aba_destino.cell(row=linha_destino, column=coluna)
        celula_destino.value = celula_origem.value


arquivo_bairros = load_workbook("AutoExcel/Bairros.xlsx")

aba_basedados = arquivo_bairros["Base de Dados"]

ultima_linha = aba_basedados.max_row

estilo_cabeçalho = copy(aba_basedados["A1"]._style)

for linha in range(2, ultima_linha + 1):
    bairro = aba_basedados.cell(row=linha, column=3).value
    if not bairro:
        break

    criar_aba(bairro, arquivo_bairros, estilo_cabeçalho)
    
    aba_destino = arquivo_bairros[bairro]
    transferir_info_aba(aba_basedados, aba_destino, linha)


arquivo_bairros.save("AutoExcel/Bairros2.xlsx")